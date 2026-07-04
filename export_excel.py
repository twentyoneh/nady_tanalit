"""
Выгрузка статистики из SQLite (stats.db) в Excel (.xlsx).

По умолчанию в листе «Метрики постов» — ОДИН свежий ряд на пост
(последний снимок). Пустые листы (Подписчики/Посты/Динамика) добавляются
только если по ним реально есть данные (их наполняет bot.py по Bot API).

Запуск:
    python export_excel.py                 # -> stats_export.xlsx (свежие данные)
    python export_excel.py --all           # включить ВСЕ снимки (история/динамика)
    python export_excel.py myfile.xlsx     # своё имя файла
"""

import sqlite3
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DB_PATH = "stats.db"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")


def _style_header(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None),
                    default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(width + 2, 60)


def _table_has_rows(con, name: str) -> bool:
    q = "SELECT name FROM sqlite_master WHERE type='table' AND name=?"
    if con.execute(q, (name,)).fetchone() is None:
        return False
    return con.execute(f"SELECT 1 FROM {name} LIMIT 1").fetchone() is not None


def export(db_path: str = DB_PATH, out_path: str = "stats_export.xlsx",
           show_all: bool = False) -> str:
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    wb = Workbook()
    wb.remove(wb.active)                      # убираем дефолтный пустой лист

    # === Метрики постов (главный лист) =====================================
    if _table_has_rows(con, "post_metrics"):
        cols = [r[1] for r in con.execute("PRAGMA table_info(post_metrics)")]
        link_sql = "link" if "link" in cols else "NULL AS link"
        base = ("SELECT channel, message_id, views, forwards, reactions_total, "
                f"reactions_json, comments, {link_sql}, taken_at "
                "FROM post_metrics pm ")
        if show_all:
            query = base + "ORDER BY channel, message_id, taken_at"
        else:
            # только последний снимок каждого поста -> один ряд на пост
            query = base + (
                "WHERE taken_at = (SELECT MAX(taken_at) FROM post_metrics x "
                "WHERE x.channel = pm.channel AND x.message_id = pm.message_id) "
                "ORDER BY channel, message_id")

        ws = wb.create_sheet("Метрики постов")
        ws.append(["Канал", "ID поста", "Просмотры", "Пересылки",
                   "Реакции", "Реакции (разбивка)", "Комментарии",
                   "Ссылка", "Снято (UTC)"])
        for m in con.execute(query).fetchall():
            ws.append([m["channel"], m["message_id"], m["views"], m["forwards"],
                       m["reactions_total"], m["reactions_json"],
                       m["comments"] if m["comments"] is not None else 0,
                       m["link"], m["taken_at"]])
            if m["link"]:
                cell = ws.cell(row=ws.max_row, column=8)
                cell.hyperlink = m["link"]
                cell.style = "Hyperlink"
        _style_header(ws)
        _autosize(ws)

    # === Подписчики / Динамика (только если есть данные Bot API) ============
    if _table_has_rows(con, "subscribers"):
        rows = con.execute(
            "SELECT channel, members, taken_at FROM subscribers "
            "ORDER BY channel, taken_at").fetchall()
        ws = wb.create_sheet("Подписчики")
        ws.append(["Канал", "Подписчиков", "Снято (UTC)"])
        for r in rows:
            ws.append([r["channel"], r["members"], r["taken_at"]])
        _style_header(ws)
        _autosize(ws)

        ws2 = wb.create_sheet("Динамика")
        ws2.append(["Канал", "Снято (UTC)", "Подписчиков", "Прирост"])
        prev = {}
        for r in rows:
            ch = r["channel"]
            delta = r["members"] - prev[ch] if ch in prev else 0
            ws2.append([ch, r["taken_at"], r["members"], delta])
            prev[ch] = r["members"]
        _style_header(ws2)
        _autosize(ws2)

    # === Посты (лог bot.py) ================================================
    if _table_has_rows(con, "posts"):
        ws3 = wb.create_sheet("Посты")
        ws3.append(["Канал", "ID сообщения", "Медиа", "Текст", "Опубликовано"])
        for p in con.execute(
                "SELECT channel, message_id, has_media, text, posted_at "
                "FROM posts ORDER BY channel, message_id").fetchall():
            ws3.append([p["channel"], p["message_id"],
                        "да" if p["has_media"] else "нет",
                        (p["text"] or "")[:200], p["posted_at"]])
        _style_header(ws3)
        _autosize(ws3)

    if not wb.sheetnames:                      # совсем нет данных
        ws = wb.create_sheet("Нет данных")
        ws.append(["В базе пока нет данных — сначала запустите сбор."])

    con.close()
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--all"]
    show_all = "--all" in sys.argv
    out = args[0] if args else "stats_export.xlsx"
    path = export(out_path=out, show_all=show_all)
    mode = "все снимки" if show_all else "по свежему снимку на пост"
    print(f"[{datetime.now():%H:%M:%S}] Готово ({mode}) -> {path}")
